import openpyxl


class CustomerData:

    test_SignupData = [{"firstName": "first", "lastName": "last", "street": "begur", "city": "bengaluru", "state": "karnataka", "zip": "56485", "phone": "56464466", "SSN":"464646464", "username": "first17", "password": "Welcome"},{"firstName": "first", "lastName": "last", "street": "begur", "city": "bengaluru", "state": "karnataka", "zip": "56485", "phone": "56464466", "SSN":"464646464", "username": "first18", "password": "Welcome"}]

    test_LoginData = [{"username":"first20", "password":"Welcome"}, {"username":"first18", "password":"Welcome"}]

    @staticmethod
    def getTestDataSignup(test_case_name):
        Dict = {}
        filepath = "D:/selenium/testData/parabankTestData_signup.xlsx"
        book = openpyxl.load_workbook(filepath)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]

    @staticmethod
    def getTestDataLogin(test_case_name):
        Dict = {}
        filepath = "D:/selenium/testData/parabankTestData_login.xlsx"
        book = openpyxl.load_workbook(filepath)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]